import csv
from io import StringIO

from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.contrib.auth.models import User
from django.db.models import Count, Sum, Q
from django.http import HttpResponse, HttpResponseForbidden, JsonResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.urls import reverse
from django.utils import timezone

from .forms import ProductForm, ReceiveStockForm, StoreRequestForm, StoreRequestItemForm
from .models import StoreProduct, StoreBatch, StoreRequest, StoreRequestItem, StoreAuditLog
from .notifications import notify_users, MANAGER_GROUPS, STORE_MANAGER_GROUPS
from .permissions import can_enter_medical_store, can_see_stock, is_store_manager, is_hospital_manager, is_admin_user
from .services import issue_request


SYSTEM_TITLE = "Medical Store System"
SYSTEM_SUBTITLE = "Medical Consumables & Supplies Workflow"


def _ctx(extra=None):
    base = {"system_title_override": SYSTEM_TITLE, "system_subtitle_override": SYSTEM_SUBTITLE}
    if extra:
        base.update(extra)
    return base


def _deny_if_no_access(user):
    return not can_enter_medical_store(user)


def _bar_rows(rows, label_key="label", value_key="value", limit=8):
    rows = list(rows)[:limit]
    max_value = max([int(r.get(value_key) or 0) for r in rows] or [0])
    result = []
    for row in rows:
        value = int(row.get(value_key) or 0)
        result.append({
            "label": row.get(label_key) or "-",
            "value": value,
            "percent": int((value / max_value) * 100) if max_value else 0,
        })
    return result


def _dashboard_charts():
    top_products_qs = (
        StoreRequestItem.objects
        .filter(issued_total_quantity__gt=0)
        .values("product__name")
        .annotate(value=Sum("issued_total_quantity"))
        .order_by("-value")[:8]
    )
    top_products = _bar_rows([
        {"label": r["product__name"], "value": r["value"] or 0}
        for r in top_products_qs
    ])

    dept_qs = (
        StoreRequest.objects
        .exclude(department="")
        .values("department")
        .annotate(value=Count("id"))
        .order_by("-value")[:8]
    )
    department_requests = _bar_rows([
        {"label": r["department"], "value": r["value"] or 0}
        for r in dept_qs
    ])

    status_qs = (
        StoreRequest.objects
        .values("status")
        .annotate(value=Count("id"))
        .order_by("-value")
    )
    status_labels = dict(StoreRequest.STATUS_CHOICES)
    request_status = _bar_rows([
        {"label": status_labels.get(r["status"], r["status"]), "value": r["value"] or 0}
        for r in status_qs
    ], limit=10)

    supplier_qs = (
        StoreBatch.objects
        .exclude(supplier_name="")
        .values("supplier_name")
        .annotate(value=Count("id"))
        .order_by("-value")[:8]
    )
    suppliers = _bar_rows([
        {"label": r["supplier_name"], "value": r["value"] or 0}
        for r in supplier_qs
    ])

    return {
        "chart_top_products": top_products,
        "chart_department_requests": department_requests,
        "chart_request_status": request_status,
        "chart_suppliers": suppliers,
    }


@login_required
def entry(request):
    if _deny_if_no_access(request.user):
        return HttpResponseForbidden("You are not allowed to access Medical Store System.")
    if is_store_manager(request.user):
        return redirect("medical_store:store_dashboard")
    if is_hospital_manager(request.user) or is_admin_user(request.user):
        return redirect("medical_store:hospital_dashboard")
    return redirect("medical_store:nurse_dashboard")


@login_required
def store_dashboard(request):
    if not is_store_manager(request.user) and not is_admin_user(request.user):
        return HttpResponseForbidden("Store Manager only.")
    today = timezone.localdate()
    near_expiry_date = today + timezone.timedelta(days=60)
    products = StoreProduct.objects.all()
    batches = StoreBatch.objects.all()
    context = {
        "total_products": products.count(),
        "total_batches": batches.count(),
        "total_stock": batches.aggregate(total=Sum("current_quantity"))["total"] or 0,
        "pending_requests": StoreRequest.objects.filter(status=StoreRequest.STATUS_APPROVED).count(),
        "issued_today": StoreRequest.objects.filter(store_issued_at__date=today).count(),
        "received_today": StoreBatch.objects.filter(received_date=today).count(),
        "low_stock_count": sum(1 for p in products if p.is_low_stock),
        "near_expiry_count": batches.filter(expiry_date__isnull=False, expiry_date__lte=near_expiry_date).count(),
        "activity": StoreAuditLog.objects.select_related("user", "store_request", "product")[:20],
    }
    context.update(_dashboard_charts())
    return render(request, "medical_store/store_dashboard.html", _ctx(context))


@login_required
def nurse_dashboard(request):
    if _deny_if_no_access(request.user):
        return HttpResponseForbidden("Access denied.")
    qs = StoreRequest.objects.filter(requested_by=request.user)
    context = {
        "draft_count": qs.filter(status=StoreRequest.STATUS_DRAFT).count(),
        "pending_count": qs.filter(status=StoreRequest.STATUS_PENDING_HM).count(),
        "approved_count": qs.filter(status=StoreRequest.STATUS_APPROVED).count(),
        "waiting_confirm_count": qs.filter(status=StoreRequest.STATUS_ISSUED).count(),
        "closed_count": qs.filter(status=StoreRequest.STATUS_CLOSED).count(),
        "rejected_count": qs.filter(status=StoreRequest.STATUS_REJECTED).count(),
    }
    return render(request, "medical_store/nurse_dashboard.html", _ctx(context))


@login_required
def hospital_dashboard(request):
    if not (is_hospital_manager(request.user) or is_admin_user(request.user)):
        return HttpResponseForbidden("Hospital Manager only.")
    today = timezone.localdate()
    near_expiry_date = today + timezone.timedelta(days=60)
    products = StoreProduct.objects.all()
    context = {
        "pending_approval": StoreRequest.objects.filter(status=StoreRequest.STATUS_PENDING_HM).count(),
        "approved": StoreRequest.objects.filter(status=StoreRequest.STATUS_APPROVED).count(),
        "waiting_store": StoreRequest.objects.filter(status=StoreRequest.STATUS_APPROVED).count(),
        "waiting_nurse": StoreRequest.objects.filter(status=StoreRequest.STATUS_ISSUED).count(),
        "closed": StoreRequest.objects.filter(status=StoreRequest.STATUS_CLOSED).count(),
        "rejected": StoreRequest.objects.filter(status=StoreRequest.STATUS_REJECTED).count(),
        "low_stock_count": sum(1 for p in products if p.is_low_stock),
        "near_expiry_count": StoreBatch.objects.filter(expiry_date__isnull=False, expiry_date__lte=near_expiry_date).count(),
    }
    context.update(_dashboard_charts())
    return render(request, "medical_store/hospital_dashboard.html", _ctx(context))


@login_required
def add_product(request):
    if not is_store_manager(request.user) and not is_admin_user(request.user):
        return HttpResponseForbidden("Store Manager only.")
    form = ProductForm(request.POST or None)
    if request.method == "POST" and form.is_valid():
        product = form.save()
        StoreAuditLog.objects.create(action="ADD_PRODUCT", user=request.user, product=product, details=f"Added/updated product {product.name}")
        messages.success(request, "Product saved successfully.")
        return redirect("medical_store:inventory")
    return render(request, "medical_store/product_form.html", _ctx({"form": form}))


@login_required
def receive_stock(request):
    if not is_store_manager(request.user) and not is_admin_user(request.user):
        return HttpResponseForbidden("Store Manager only.")
    form = ReceiveStockForm(request.POST or None)
    if request.method == "POST" and form.is_valid():
        batch = form.save(commit=False)
        batch.received_unit = form.cleaned_data.get("received_unit") or batch.product.unit
        batch.received_unit_pieces = batch.product.conversion_for(batch.received_unit)
        batch.received_total_quantity = batch.received_quantity * batch.received_unit_pieces
        batch.current_quantity = batch.received_total_quantity
        batch.received_by = request.user
        batch.save()
        StoreAuditLog.objects.create(action="RECEIVE_STOCK", user=request.user, product=batch.product, details=f"Received {batch.received_display} = {batch.received_total_quantity} base units - Batch {batch.batch_number or 'N/A'}")
        messages.success(request, "Stock received and added to inventory.")
        return redirect("medical_store:inventory")
    product_units = {str(p.id): {"units": p.get_allowed_units(), "unit_pieces": p.unit_pieces or {}} for p in StoreProduct.objects.filter(is_active=True)}
    return render(request, "medical_store/receive_stock.html", _ctx({"form": form, "product_units": product_units}))


@login_required
def inventory(request):
    if not can_see_stock(request.user):
        return HttpResponseForbidden("You are not allowed to view stock balance.")
    q = request.GET.get("q", "").strip()
    products = StoreProduct.objects.all()
    if q:
        products = products.filter(Q(name__icontains=q) | Q(code__icontains=q) | Q(manufacturer__icontains=q) | Q(shelf__icontains=q))
    return render(request, "medical_store/inventory.html", _ctx({"products": products, "q": q}))


@login_required
def product_detail(request, product_id):
    if not can_see_stock(request.user):
        return HttpResponseForbidden("You are not allowed to view stock balance.")
    product = get_object_or_404(StoreProduct, id=product_id)
    return render(request, "medical_store/product_detail.html", _ctx({"product": product}))


def _department_for_user(user):
    groups = set(user.groups.values_list("name", flat=True))
    mapping = [
        (["MALE"], "Male Clinics"),
        (["FEMALE"], "Female Clinics"),
        (["SPECIALTY"], "Specialty Clinics"),
        (["EMERGENCY", "EMR"], "Emergency Clinics"),
        (["CSSD"], "CSSD"),
        (["RADIOLOGY"], "Radiology"),
        (["LAB"], "Laboratory"),
    ]
    for keys, label in mapping:
        if any(k in groups for k in keys):
            return label
    return user.get_full_name() or user.username


@login_required
def product_search(request):
    if _deny_if_no_access(request.user):
        return JsonResponse({"results": []})
    q = request.GET.get("q", "").strip()
    products = StoreProduct.objects.filter(is_active=True)
    if q:
        products = products.filter(Q(name__icontains=q) | Q(code__icontains=q))
    products = products.order_by("name")[:20]
    results = []
    for p in products:
        units = p.get_allowed_units()
        results.append({
            "id": p.id,
            "name": p.name,
            "code": p.code,
            "label": f"{p.name} - {p.code}",
            "units": units,
            "unit_pieces": p.unit_pieces or {},
        })
    return JsonResponse({"results": results})


@login_required
def new_request(request):
    if _deny_if_no_access(request.user):
        return HttpResponseForbidden("Access denied.")
    form = StoreRequestForm(request.POST or None)
    department = _department_for_user(request.user)

    if request.method == "POST" and form.is_valid():
        action = request.POST.get("action", "submit")
        store_request = form.save(commit=False)
        store_request.requested_by = request.user
        store_request.department = department
        store_request.status = StoreRequest.STATUS_DRAFT if action == "draft" else StoreRequest.STATUS_PENDING_HM
        store_request.save()
        valid_items = 0
        row_indexes = request.POST.getlist("row_index")
        for row in row_indexes:
            product_id = request.POST.get(f"product_{row}")
            qty = request.POST.get(f"quantity_{row}")
            unit = request.POST.get(f"unit_{row}")
            if not product_id or not qty or not unit:
                continue
            try:
                qty = int(qty)
            except ValueError:
                qty = 0
            if qty <= 0:
                continue
            product = StoreProduct.objects.filter(id=product_id, is_active=True).first()
            if product and unit in product.get_allowed_units():
                unit_pieces = product.conversion_for(unit)
                StoreRequestItem.objects.create(
                    store_request=store_request,
                    product=product,
                    requested_quantity=qty,
                    requested_unit=unit,
                    requested_unit_pieces=unit_pieces,
                    requested_total_quantity=qty * unit_pieces,
                )
                valid_items += 1
        if valid_items == 0:
            store_request.delete()
            messages.error(request, "Please add at least one item.")
            return redirect("medical_store:new_request")
        StoreAuditLog.objects.create(action="CREATE_REQUEST" if action != "draft" else "SAVE_DRAFT", user=request.user, store_request=store_request, details=store_request.reason)
        if action != "draft":
            notify_users(
                title="Medical Store Request Waiting Approval",
                message=f"{store_request.request_number} is waiting for hospital manager approval.",
                url=reverse("medical_store:request_detail", args=[store_request.id]),
                groups=MANAGER_GROUPS,
            )
        messages.success(request, "Request saved successfully.")
        return redirect("medical_store:request_detail", request_id=store_request.id)
    return render(request, "medical_store/request_form.html", _ctx({"form": form, "department": department}))


@login_required
def request_list(request, status="all"):
    if is_hospital_manager(request.user) or is_admin_user(request.user):
        qs = StoreRequest.objects.all()
    elif is_store_manager(request.user):
        qs = StoreRequest.objects.exclude(status=StoreRequest.STATUS_DRAFT)
    else:
        qs = StoreRequest.objects.filter(requested_by=request.user)
    if status != "all":
        qs = qs.filter(status=status)
    return render(request, "medical_store/request_list.html", _ctx({"requests": qs, "status": status}))


@login_required
def request_detail(request, request_id):
    store_request = get_object_or_404(StoreRequest, id=request_id)
    allowed = (
        is_admin_user(request.user) or is_hospital_manager(request.user) or is_store_manager(request.user)
        or store_request.requested_by_id == request.user.id
    )
    if not allowed:
        return HttpResponseForbidden("Access denied.")
    return render(request, "medical_store/request_detail.html", _ctx({"store_request": store_request}))


@login_required
def approve_request(request, request_id):
    if not (is_hospital_manager(request.user) or is_admin_user(request.user)):
        return HttpResponseForbidden("Hospital Manager only.")
    store_request = get_object_or_404(StoreRequest, id=request_id, status=StoreRequest.STATUS_PENDING_HM)
    if request.method == "POST":
        action = request.POST.get("action")
        comment = request.POST.get("comment", "")
        if action == "reject":
            store_request.status = StoreRequest.STATUS_REJECTED
            store_request.hospital_manager_comment = comment
            msg = "rejected"
        else:
            store_request.status = StoreRequest.STATUS_APPROVED
            store_request.hospital_manager_comment = comment
            msg = "approved"
        store_request.hospital_manager_by = request.user
        store_request.hospital_manager_at = timezone.now()
        store_request.save()
        StoreAuditLog.objects.create(action=f"HOSPITAL_MANAGER_{msg.upper()}", user=request.user, store_request=store_request, details=comment)
        if store_request.status == StoreRequest.STATUS_APPROVED:
            notify_users("Medical Store Request Approved", f"{store_request.request_number} is approved and waiting store issue.", reverse("medical_store:request_detail", args=[store_request.id]), groups=STORE_MANAGER_GROUPS)
        else:
            notify_users("Medical Store Request Rejected", f"{store_request.request_number} was rejected.", reverse("medical_store:request_detail", args=[store_request.id]), users=[store_request.requested_by])
        return redirect("medical_store:request_detail", request_id=store_request.id)
    return render(request, "medical_store/approve_request.html", _ctx({"store_request": store_request}))


@login_required
def issue_request_view(request, request_id):
    if not (is_store_manager(request.user) or is_admin_user(request.user)):
        return HttpResponseForbidden("Store Manager only.")
    store_request = get_object_or_404(StoreRequest, id=request_id, status=StoreRequest.STATUS_APPROVED)
    if request.method == "POST":
        issue_quantities = {str(item.id): request.POST.get(f"issued_{item.id}", 0) for item in store_request.items.all()}
        issued_units = {str(item.id): request.POST.get(f"issued_unit_{item.id}", item.requested_unit) for item in store_request.items.all()}
        reasons = {str(item.id): request.POST.get(f"reason_{item.id}", "") for item in store_request.items.all()}
        comment = request.POST.get("comment", "")
        issue_request(store_request, request.user, issue_quantities, reasons, comment, issued_units)
        notify_users("Medical Store Items Issued", f"{store_request.request_number} has been issued. Please confirm receipt.", reverse("medical_store:request_detail", args=[store_request.id]), users=[store_request.requested_by])
        messages.success(request, "Items issued successfully.")
        return redirect("medical_store:request_detail", request_id=store_request.id)
    return render(request, "medical_store/issue_request.html", _ctx({"store_request": store_request}))


@login_required
def confirm_receipt(request, request_id):
    store_request = get_object_or_404(StoreRequest, id=request_id, status=StoreRequest.STATUS_ISSUED)
    if store_request.requested_by_id != request.user.id and not is_admin_user(request.user):
        return HttpResponseForbidden("Requester only.")
    if request.method == "POST":
        store_request.status = StoreRequest.STATUS_CLOSED
        store_request.nurse_confirmed_by = request.user
        store_request.nurse_confirmed_at = timezone.now()
        store_request.nurse_comment = request.POST.get("comment", "")
        store_request.save()
        StoreAuditLog.objects.create(action="NURSE_CONFIRMED_RECEIPT", user=request.user, store_request=store_request, details=store_request.nurse_comment)
        notify_users("Medical Store Request Closed", f"{store_request.request_number} has been confirmed and closed.", reverse("medical_store:request_detail", args=[store_request.id]), groups=STORE_MANAGER_GROUPS + MANAGER_GROUPS)
        messages.success(request, "Receipt confirmed. Request closed.")
        return redirect("medical_store:request_detail", request_id=store_request.id)
    return render(request, "medical_store/confirm_receipt.html", _ctx({"store_request": store_request}))


@login_required
def low_stock(request):
    if not can_see_stock(request.user):
        return HttpResponseForbidden("Access denied.")
    products = [p for p in StoreProduct.objects.all() if p.is_low_stock]
    return render(request, "medical_store/low_stock.html", _ctx({"products": products}))


@login_required
def near_expiry(request):
    if not can_see_stock(request.user):
        return HttpResponseForbidden("Access denied.")
    today = timezone.localdate()
    batches = StoreBatch.objects.filter(expiry_date__isnull=False, expiry_date__lte=today + timezone.timedelta(days=60)).select_related("product")
    return render(request, "medical_store/near_expiry.html", _ctx({"batches": batches}))


@login_required
def reports_center(request):
    if not can_see_stock(request.user):
        return HttpResponseForbidden("Access denied.")
    departments = StoreRequest.objects.exclude(department="").values("department").annotate(total=Count("id")).order_by("department")
    return render(request, "medical_store/reports_center.html", _ctx({"departments": departments}))


@login_required
def department_requests_report(request):
    if not can_see_stock(request.user):
        return HttpResponseForbidden("Access denied.")
    departments = StoreRequest.objects.exclude(department="").values_list("department", flat=True).distinct().order_by("department")
    selected_department = request.GET.get("department", "").strip()
    status = request.GET.get("status", "").strip()
    date_from = request.GET.get("from", "").strip()
    date_to = request.GET.get("to", "").strip()
    qs = StoreRequest.objects.select_related("requested_by").all()
    if selected_department:
        qs = qs.filter(department=selected_department)
    if status:
        qs = qs.filter(status=status)
    if date_from:
        qs = qs.filter(created_at__date__gte=date_from)
    if date_to:
        qs = qs.filter(created_at__date__lte=date_to)
    return render(request, "medical_store/department_requests_report.html", _ctx({
        "requests": qs,
        "departments": departments,
        "selected_department": selected_department,
        "selected_status": status,
        "date_from": date_from,
        "date_to": date_to,
        "status_choices": StoreRequest.STATUS_CHOICES,
    }))


@login_required
def export_inventory_csv(request):
    if not can_see_stock(request.user):
        return HttpResponseForbidden("Access denied.")
    response = HttpResponse(content_type="text/csv")
    response["Content-Disposition"] = 'attachment; filename="medical_store_inventory.csv"'
    writer = csv.writer(response)
    writer.writerow(["Product", "Code", "Manufacturer", "Allowed Units", "Shelf", "Minimum Stock", "Available", "Display Stock"])
    for p in StoreProduct.objects.all():
        writer.writerow([p.name, p.code, p.manufacturer, ", ".join(p.get_allowed_units()), p.shelf, p.minimum_stock, p.total_available, p.display_stock])
    return response


@login_required
def import_products(request):
    if not (is_store_manager(request.user) or is_admin_user(request.user)):
        return HttpResponseForbidden("Store Manager only.")
    if request.method == "POST" and request.FILES.get("file"):
        uploaded = request.FILES["file"]
        count = 0
        if uploaded.name.lower().endswith(".csv"):
            data = uploaded.read().decode("utf-8-sig")
            rows = csv.DictReader(StringIO(data))
            for row in rows:
                code = (row.get("code") or row.get("Code") or "").strip()
                name = (row.get("name") or row.get("Name") or row.get("product") or row.get("Product") or "").strip()
                if not code or not name:
                    continue
                StoreProduct.objects.update_or_create(
                    code=code,
                    defaults={
                        "name": name,
                        "manufacturer": row.get("manufacturer") or row.get("Manufacturer") or "",
                        "unit": row.get("unit") or row.get("Unit") or "Piece",
                        "allowed_units": [row.get("unit") or row.get("Unit") or "Piece"],
                        "unit_pieces": {"Pack": int(row.get("pack_pieces") or 1), "Carton": int(row.get("carton_pieces") or 1), "Set": int(row.get("set_pieces") or 1)},
                        "shelf": row.get("shelf") or row.get("Shelf") or "",
                        "minimum_stock": int(row.get("minimum_stock") or row.get("Minimum Stock") or 0),
                        "category": row.get("category") or row.get("Category") or "",
                    },
                )
                count += 1
        else:
            try:
                from openpyxl import load_workbook
                wb = load_workbook(uploaded, data_only=True)
                ws = wb.active
                headers = [str(c.value).strip().lower().replace(" ", "_") if c.value else "" for c in ws[1]]
                for row in ws.iter_rows(min_row=2, values_only=True):
                    data = dict(zip(headers, row))
                    code = str(data.get("code") or "").strip()
                    name = str(data.get("name") or data.get("product") or "").strip()
                    if not code or not name:
                        continue
                    StoreProduct.objects.update_or_create(
                        code=code,
                        defaults={
                            "name": name,
                            "manufacturer": data.get("manufacturer") or "",
                            "unit": data.get("unit") or "Piece",
                            "allowed_units": [data.get("unit") or "Piece"],
                            "unit_pieces": {"Pack": int(data.get("pack_pieces") or 1), "Carton": int(data.get("carton_pieces") or 1), "Set": int(data.get("set_pieces") or 1)},
                            "shelf": data.get("shelf") or "",
                            "minimum_stock": int(data.get("minimum_stock") or 0),
                            "category": data.get("category") or "",
                        },
                    )
                    count += 1
            except Exception as exc:
                messages.error(request, f"Excel import failed: {exc}. You can upload CSV instead.")
                return redirect("medical_store:import_products")
        StoreAuditLog.objects.create(action="IMPORT_PRODUCTS", user=request.user, details=f"Imported {count} products")
        messages.success(request, f"Imported/updated {count} products.")
        return redirect("medical_store:inventory")
    return render(request, "medical_store/import_products.html", _ctx())
